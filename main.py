#Libraries
from datetime import date
from datetime import datetime
import os
import openpyxl
from openpyxl import Workbook
import openpyxl.workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
import re

# --Functions--
# Excel File Creation
def file_creation(path):
  crntfile = Workbook()
  crntfile.save(path)
  print('A new excel sheet has been created.')
  return crntfile

# Excel File Appending
def file_append(path):
  crntfile = openpyxl.load_workbook(path)
  print('The existing excel sheet will now be modified.')
  file_exist = True
  return crntfile

# Add a record
def add_record():
  input_list = []
  time_input_dec = input('Would you like to enter time of income/expense (T) or use Current Time (C): ')
  while time_input_dec.upper() != 'T' and time_input_dec.upper() != 'C':
    print('Wrong Input! ')
    time_input_dec = input('Would you like to enter time of income/expense (T) or use Current Time (C): ')
  if time_input_dec.upper() == 'T': #Inputting time
    time_input = input('Enter the time of income/expense (HH:MM): ')
    while not re.match(time_pattern, time_input):
      print('Wrong Format! ')
      time_input = input('Enter the time of income/expense (HH:MM): ')
  elif time_input_dec.upper() == 'C':
    current_datetime = datetime.now()
    time_input = current_datetime.strftime("%H:%M")
  input_list.append(time_input)
  for i in range(len(headertitles)-2):
    user_input = input(f'Enter the {headertitles[i+1]} if applicable: ')
    if i in [2,3]:
      y = False 
      while y == False:
        try:
          user_input = float(user_input)
          y = True
        except:
          print('Enter a number!')
          user_input = input(f'Enter the {headertitles[i+1]} if applicable: ')
    input_list.append(user_input)
  input_tuple = (input_list)
  ws.append(input_tuple)
  currentfile.save(filepath)
  
# Income, Expense, Balance stats store
def stats():
  Total_Income = 0
  for i in range(2, ws.max_row):
    Total_Income = Total_Income + float(ws.cell(row=i, column = 4).value)
    ws.cell(row=ws.max_row, column = 4, value=Total_Income)
  Total_Expense = 0
  for i in range(2, ws.max_row):
    Total_Expense = Total_Expense + float(ws.cell(row=i, column = 5).value)
    ws.cell(row=ws.max_row, column = 5, value=Total_Expense)
  Net_Balance = float(ws.cell(row=ws.max_row-1, column = ws.max_column).value)
  ws.cell(row=ws.max_row, column = 6, value=Net_Balance)

#Border preset 
borders = Side(border_style="thin", color="000000")

#Validation patterns
time_pattern = r'(\d\d):(\d\d)'
date_pattern = r'(\d\d\d\d)-(\d\d)-(\d\d)'

#Title
Titlestr = 'DAILY EXPENSE TRACKER'
trgtdate = date.today()
print(f"\n{Titlestr:-^70} \nToday's Date: {trgtdate}\n")

#Optional Date Input
answer = True
while answer == True:
  trgtdate = str(trgtdate)
  diffdate = input("Would you like to manage or track income/expenses for a different date (Y/N)? ")
  while diffdate.upper() != 'Y' and diffdate.upper() != 'N':
    print('Enter Y or N! ')
    diffdate = input("Would you like to manage or track income/expenses for a different date (Y/N)? ")
  if diffdate.upper() == 'Y':
    trgtdate = input('Enter target date in YYYY-MM-DD format (Include \'-\'): ')
  while not re.match(date_pattern, trgtdate):
      print('Wrong Format! ')
      trgtdate = input('Enter target date in YYYY-MM-DD format (Include \'-\'): ')
      
  # File Existence Check
  filename = f'{trgtdate}-expenses.xlsx'
  foldername = 'Daily-Income-Expense-Sheets'
  filepath = os.path.join(foldername, filename)
  file_exist = False
  if os.path.exists(filepath) and os.path.isfile(filepath):
    currentfile = file_append(filepath) # Append file since file exists
  else:
    currentfile = file_creation(filepath) # Create file since file does not exist

  # Default Header Writing
  ws = currentfile.active
  headertitles = ['Time', 'Description', 'Category', 'Income', 'Expense', 'Balance']
  if file_exist is False:
    for i in range(6):
      ws.cell(row=1, column = i+1, value = headertitles[i])
      ws.cell(row=1, column = i+1).font = Font(bold=True)
      ws.cell(row=1, column = i+1).fill = PatternFill('solid', start_color="808080")
      ws.cell(row=1, column = i+1).border = Border(
        top = borders, left = borders, right = borders, bottom = borders
      )
    currentfile.save(filepath)
    
  # Record
  recordask = input('Would you like to enter a new record (Y/N)? ')
  while recordask.upper() != 'Y' and recordask.upper() != 'N':
    print('Enter Y or N! ')
    recordask = input('Would you like to enter a new record (Y/N)? ')
  while recordask.upper() == 'Y':
    if ws.max_row != 1:
      ws.delete_rows(ws.max_row, 1)
      currentfile.save(filepath)
    add_record()
    if ws.max_row==2:
      ws.cell(row=ws.max_row, column=6, value=(float(ws.cell(row=ws.max_row, column=4).value) - float(ws.cell(row=ws.max_row, column=5).value)))
    else:
      ws.cell(row=ws.max_row, column=6, value=(float(ws.cell(row=ws.max_row, column=4).value) - float(ws.cell(row=ws.max_row, column=5).value) + float(ws.cell(row=ws.max_row - 1, column=6).value)))
    currentfile.save(filepath)
    ws.cell(row=ws.max_row+1, column = 1, value = 'Total: ')
    currentfile.save(filepath)
    stats()
    currentfile.save(filepath)
    recordask = input('Would you like to enter a new record (Y/N)? ')
    while recordask.upper() != 'Y' and recordask.upper() != 'N':
      print('Enter Y or N! ')
      recordask = input('Would you like to enter a new record (Y/N)? ')

  # Displaying total income, total expense, Net Balance
  statsask = input('Would you like to view the total income, total expenses and net balance for this day (Y/N)? ')
  while statsask.upper() != 'Y' and statsask.upper() != 'N':
    print('Enter Y or N! ')
    statsask = input('Would you like to view the total income, total expenses and net balance for this day (Y/N)? ')
     
  if statsask.upper() == 'Y':
    print(f'As of now, the total income is {float(ws.cell(row=ws.max_row, column=4).value)}, the total expense is {float(ws.cell(row=ws.max_row, column=5).value)} and the total balance is {float(ws.cell(row=ws.max_row, column=6).value)}')

  user_answer = input('Would you like to exit (Y/N)? ')
  while user_answer.upper() != 'Y' and user_answer.upper() != 'N':
    print('Enter Y or N! ')
    user_answer = input('Would you like to exit (Y/N)? ')
  if user_answer.upper() == 'Y':
    answer = False
  else:
    print(f"\nToday's Date: {trgtdate}\n")

#Ending Message
print('The application has closed.')